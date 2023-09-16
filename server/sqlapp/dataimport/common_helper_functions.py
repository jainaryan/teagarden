import openpyxl
from sqlapp.models import *
from sqlapp.database import db_session, init_db


def find_year(sheet):
    row = 4
    for column in range(1,14):
        if sheet.cell(row=row, column=column).value == None:
            pass
        else:
            return row, column


def count_stations_for_entity(entity_id: int) -> int:
    return db_session.query(Station).filter_by(entity_id=entity_id).count()

# Function to get the station ID for an entity if there's only one station
def get_station_id_for_entity(entity_id: int) -> int:
    station = db_session.query(Station).filter_by(entity_id=entity_id).first()
    if station:
        return station.id
    else:
        # Handle the case where no station is found
        return None



def geoEntity_entry(name: str, area=None, latitude=None, longitude=None, district=None, state='Assam'):
    entity = GeoEntity(name=name, latitude=latitude, longitude=longitude, district=district, state=state, area=area)
    db_session.add(entity)
    db_session.commit()


def get_entity_id(entityName: str):
    entity = db_session.query(GeoEntity).filter_by(name=entityName).first()
    return entity.entity_id


def get_sensor(stationName: str):
    sensor = db_session.query(Station).filter_by(sensor_name=stationName).first()
    return sensor


def read_values(sheet):
    name = sheet.cell(row=1, column=2).value
    entity_id = sheet.cell(row=2, column=2).value
    station_id = sheet.cell(row=3, column=2).value
    latitude = sheet.cell(row=4, column=2).value
    longitude = sheet.cell(row=5, column=2).value
    district = sheet.cell(row=6, column=2).value
    state = sheet.cell(row=7, column=2).value
    area = sheet.cell(row=8, column=2).value
    unit = sheet.cell(row=10, column=2).value
    return name, entity_id, station_id, latitude, longitude, district, state, area, unit

def check_entity_name(entityName: str):
    entity = (db_session.query(GeoEntity).filter_by(name=entityName).first())

    if entity == None:
        # garden is not present in table
        return False
    else:
        # garden is present in table
        return True

def get_year(year):
    temp_year = year.split()
    temp_year = temp_year[1]
    return int(temp_year)

def create_station_entry(entity_id: int):
    stationName = None
    station_entry(entity_id=entity_id, name=stationName)
    station_id = (db_session.query(Station).filter_by(entity_id=entity_id).first()).id
    return station_id


def station_entry(entity_id: int, name=None, lat=None, long=None, type=None):
    init_db()
    if name == None:
        entity = db_session.query(GeoEntity).filter_by(entity_id=entity_id).first()
        name = str(entity.name) + "_sensor" + str(entity.entity_id)
    db_session.add(Station(sensor_name=name, entity_id=entity_id, sensor_type=type, latitude=lat, longitude=long))
    db_session.commit()
