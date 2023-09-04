from datetime import datetime
import openpyxl
from sqlapp.models  import *
from sqlapp.database import *
def type_1_reader(workbook):
    first_sheet = True
    for sheet in workbook._sheets:
        if first_sheet:
            entityName, latitude, longitude, district, state, area = read_values(sheet)
            if check_entity_name(entityName):
                # entity in table
                entity_id = sheet.cell(row=2, column=2).value
                station_id = sheet.cell(row=3, column=2).value
            else:
                # entity not in table
                geoEntity_entry(name=entityName, latitude=latitude, longitude=longitude, district=district, state=state,
                                area=area)
                entity_id = get_entity_id(entityName)
                station_id = create_station_entry(entity_id)
            first_sheet = False
        else:
            format_sheet(sheet, workbook)
            input_data_from_excel(sheet, station_id)


def format_sheet(sheet, workbook):

    sheet.cell(row=1, column=7).value = ""
    sheet.cell(row=1, column=1).value = sheet.cell(row=4, column=7).value
    for column in range(1, 14):
        for row in range(2, 37):
            sheet.cell(row=row, column=column).value = sheet.cell(row=row + 3, column=column).value
    #there will be a problem here as multiple files will be stored with the same name
    workbook.save("done.xlsx")



def input_data_from_excel(sheet, station_id):
    temp_year = sheet.cell(row=1, column=1).value
    year = get_year(temp_year)
    first_month = 2
    last_month = 13
    first_day = 3
    for month in range(first_month, last_month + 1):
        last_day = find_last_day_of_month(month, int(year)) + 2
        for day in range(first_day, last_day + 1):
            check_reading = (sheet.cell(row=day, column=month).value)
            if check_reading == 'N/A' or check_reading == None:
                break
            else:
                reading = float(sheet.cell(row=day, column=month).value)
                date = convert_date(year, month - 1, day - 2)
                rainfallData_entry(station_id=station_id, reading=reading, date=date)


def find_last_day_of_month(month: int, year: int):
    if month == 3:
        if year % 4 == 0:
            return 29
        else:
            return 28

    elif (month < 9):
        if (month % 2 == 0):
            return 31
        else:
            return 30

    elif (month > 9):
        if (month % 2 == 0):
            return 30
        else:
            return 31
    else:
        return 31





def read_values(sheet):
    name = sheet.cell(row=1, column=2).value
    latitude = sheet.cell(row=4, column=2).value
    longitude = sheet.cell(row=5, column=2).value
    district = sheet.cell(row=6, column=2).value
    state = sheet.cell(row=7, column=2).value
    area = sheet.cell(row=8, column=2).value
    return name, latitude, longitude, district, state, area


def geoEntity_entry(name: str, area=None, latitude=None, longitude=None, district=None, state='Assam'):
    entity = GeoEntity(name=name, latitude=latitude, longitude=longitude, district=district, state=state, area=area)
    db_session.add(entity)
    db_session.commit()


def get_entity_id(entityName: str):
    entity = db_session.query(GeoEntity).filter_by(name=entityName).first()
    return entity.id


def get_sensor(stationName: str):
    sensor = db_session.query(Station).filter_by(sensor_name=stationName).first()
    return sensor


def create_station_entry(entity_id: int):
    stationName = None
    station_entry(entity_id=entity_id, name=stationName)
    station_id = (db_session.query(Station).filter_by(entity_id=entity_id).first()).id
    return station_id


def station_entry(entity_id: int, name=None, lat=None, long=None, type=None):
    init_db()
    if name == None:
        garden = db_session.query(GeoEntity).filter_by(id=entity_id).first()
        name = str(garden.name) + "_sensor" + str(garden.id)
    db_session.add(Station(sensor_name=name, entity_id=entity_id, sensor_type=type, latitude=lat, longitude=long))
    db_session.commit()


def check_entity_name(entityName: str):
    entity = (db_session.query(GeoEntity).filter_by(name=entityName).first())

    if entity == None:
        # garden is not present in table
        return False
    else:
        # garden is present in table
        return True


def rainfallData_entry(station_id: int, reading: float, date: datetime):
    rainfalldata = RainfallData(station_id=station_id, reading=reading, date=date)
    db_session.add(rainfalldata)
    db_session.commit()


def get_year(year: int):
    temp_year = year[4:]
    temp_year = temp_year.strip()
    return temp_year


def convert_date(year: int, month: int, day: int):
    temp_date = "" + str(year) + "-" + str(month) + "-" + str(day)
    converted_date = datetime.strptime(temp_date, '%Y-%m-%d').date()
    return converted_date