from datetime import datetime, time
import openpyxl
from numpy import NaN
from common_helper_functions import *
from sqlapp.models import *
from sqlapp.database import *


def rainfall_type_1_reader(workbook):
    first_sheet = True
    for sheet in workbook._sheets:
        if first_sheet:
            entityName, entity_id, station_id, latitude, longitude, district, state, area, unit = read_values(sheet)
            if check_entity_name(entityName):
                if (entity_id == None):
                    entity_id = get_entity_id(entityName)
                    station_count = count_stations_for_entity(entity_id)
                    if station_count == 1:
                        # Return the station ID
                        station_id = get_station_id_for_entity(entity_id)
            else:
                # entity not in table
                geoEntity_entry(name=entityName, latitude=latitude, longitude=longitude, district=district, state=state,
                                area=area)
                entity_id = get_entity_id(entityName)
                station_id = create_station_entry(entity_id)
            first_sheet = False
        else:
            format_sheet(sheet, workbook)
            input_data_from_excel(sheet, station_id, unit)




def format_sheet(sheet, workbook):
    #sheet.cell(row=1, column=7).value = ""
    year_row, year_column = find_year(sheet)
    sheet.cell(row=1, column=1).value = sheet.cell(row=year_row, column=year_column).value
    for column in range(1, 14):
        for row in range(2, 37):
            sheet.cell(row=row, column=column).value = sheet.cell(row=row + 3, column=column).value
            if sheet.cell(row = row, column = column).value==None:
                sheet.cell(row=row, column=column).value = 0
    # there will be a problem here as multiple files will be stored with the same name
    workbook.save("done.xlsx")


def input_data_from_excel(sheet, station_id, unit):
    if (unit == 'inches'):
        multiplier = 25.4
    elif (unit == 'millimetres'):
        multiplier = 1
    elif (unit == 'centimetres'):
        multiplier = 10
    temp_year = sheet.cell(row=1, column=1).value
    year = get_year(temp_year)
    first_month = 2
    last_month = 13
    first_day = 3
    for month in range(first_month, last_month + 1):
        last_day = find_last_day_of_month(month, int(year)) + 2
        for day in range(first_day, last_day + 1):
            check_reading = (sheet.cell(row=day, column=month).value)
            if check_reading == 'N/A':
                start_time, end_time = convert_date(year, month - 1, day - 2)
                rainfallData_entry(station_id=station_id, reading=NaN, start_time=start_time, end_time=end_time)
            else:
                reading = float(sheet.cell(row=day, column=month).value) * multiplier
                reading = round(reading,2)
                start_time, end_time = convert_date(year, month - 1, day - 2)
                rainfallData_entry(station_id=station_id, reading=reading, start_time=start_time, end_time=end_time)


def find_last_day_of_month(month: int, year: int):
    if (month == 3):
        if (year % 4) == 0:
            return 29
        else:
            return 28

    elif (month < 9):
        if (month % 2 == 0):
            return 31
        else:
            return 30

    elif (month > 9):
        if (month % 2 == 0):
            return 30
        else:
            return 31
    else:
        return 31


def rainfallData_entry(station_id: int, reading: float, start_time: datetime, end_time: datetime):
    rainfalldata = RainfallData(station_id=station_id, reading=reading, start_time=start_time, end_time=end_time)
    db_session.add(rainfalldata)
    db_session.commit()


def convert_date(year: int, month: int, day: int):
    temp_date = "" + str(year) + "-" + str(month) + "-" + str(day)
    converted_date = datetime.strptime(temp_date, '%Y-%m-%d').date()
    start_time = datetime.combine(converted_date, time(0, 0))

    # Set the end time to 11:59 pm
    end_time = datetime.combine(converted_date, time(23, 59))
    return start_time, end_time
